import streamlit as st
from streamlit_gsheets import GSheetsConnection
from datetime import datetime, date, timedelta
import pandas as pd
from streamlit_option_menu import option_menu
import plotly.express as px
import urllib.parse
import json
import io

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Border, Side, Alignment
except ImportError:
    st.error("🚨 Falta instalar openpyxl. Escribí 'pip install openpyxl' en la terminal de Cursor y agregalo a requirements.txt.")

st.set_page_config(page_title="CRM DICAD AMÉRICA", layout="wide")

USUARIOS = st.secrets["passwords"]
ADMINISTRADOR = "Ricardo Ippolito"
GSHEET_URL = st.secrets.get("SHEET_URL", "")
conn = st.connection("gsheets", type=GSheetsConnection)

CODIGOS_PAISES = [
    "🇦🇬 Antigua y Barbuda (+1)", "🇦🇷 Argentina (+54)", "🇧🇸 Bahamas (+1)", "🇧🇧 Barbados (+1)",
    "🇧🇿 Belice (+501)", "🇧🇴 Bolivia (+591)", "🇧🇷 Brasil (+55)", "🇨🇦 Canadá (+1)",
    "🇨🇱 Chile (+56)", "🇨🇴 Colombia (+57)", "🇨🇷 Costa Rica (+506)", "🇨🇺 Cuba (+53)",
    "🇩🇲 Dominica (+1)", "🇪🇨 Ecuador (+593)", "🇸🇻 El Salvador (+503)", "🇪🇸 España (+34)",
    "🇺🇸 Estados Unidos (+1)", "🇬🇩 Granada (+1)", "🇬🇹 Guatemala (+502)", "🇬🇾 Guyana (+592)",
    "🇭🇹 Haití (+509)", "🇭🇳 Honduras (+504)", "🇯🇲 Jamaica (+1)", "🇲🇽 México (+52)",
    "🇲🇿 Mozambique (+258)", "🇳🇮 Nicaragua (+505)", "🇵🇦 Panamá (+507)", "🇵🇾 Paraguay (+595)",
    "🇵🇪 Perú (+51)", "🇵🇹 Portugal (+351)", "🇩🇴 Rep. Dominicana (+1)", "🇰🇳 San Cristóbal y Nieves (+1)",
    "🇻🇨 San Vicente y las Granadinas (+1)", "🇱🇨 Santa Lucía (+1)", "🇸🇷 Surinam (+597)",
    "🇹🇹 Trinidad y Tobago (+1)", "🇺🇾 Uruguay (+598)", "🇻🇪 Venezuela (+58)", "🌎 Otro"
]

COLUMNS_MAIN = ["Cliente", "Profesion", "Direccion", "Pais", "Ciudad", "Estado /Prov.", "Empresa", "Cargo", "Telefono", "Email", "N° Cotiz.", "Monto USD / $", "Notas", "Proxima llamada", "Creado", "Asesor", "Estado_Nego", "Link_PDF", "Productos Seleccionados", "Descuento Aplicado"]
COLUMNS_CAT = ["Producto", "Descripcion", "Categoria", "Moneda", "Precio"]

def extraer_pais_codigo(seleccion):
    if seleccion == "🌎 Otro": return "Otro", ""
    try: return seleccion.split(" ", 1)[1].split(" (")[0].strip(), seleccion.split(" ", 1)[1].split(" (")[1].replace(")", "").strip()
    except: return "Desconocido", ""

def limpiar_monto_para_suma(val_str):
    texto = str(val_str).upper().replace("USD", "").replace("ARS", "").replace("$", "").strip()
    clean_str = ''.join(c for c in texto if c.isdigit() or c in '.,')
    if not clean_str: return 0.0
    if ',' in clean_str and '.' in clean_str:
        last_sep = ',' if clean_str.rfind(',') > clean_str.rfind('.') else '.'
        clean_str = clean_str.replace(',' if last_sep == '.' else '.', '').replace(last_sep, '.')
    elif ',' in clean_str: clean_str = clean_str.replace(',', '.') if len(clean_str.split(',')[-1]) != 3 else clean_str.replace(',', '')
    elif '.' in clean_str: clean_str = clean_str.replace('.', '') if len(clean_str.split('.')[-1]) == 3 else clean_str
    try: return float(clean_str)
    except: return 0.0

def parsear_fecha_hora(fecha_str):
    try:
        parts = str(fecha_str).strip().split()
        f_obj = datetime.strptime(parts[0], "%d/%m/%Y").date()
        h_obj = datetime.strptime(parts[1], "%H:%M").time() if len(parts) > 1 else datetime.strptime("10:00", "%H:%M").time()
        return f_obj, h_obj
    except: return date.today(), datetime.strptime("10:00", "%H:%M").time()

def generar_link_gcal(cliente, empresa, telefono, fecha_str):
    try:
        dt = datetime.strptime(fecha_str, "%d/%m/%Y %H:%M") if len(str(fecha_str)) > 10 else datetime.strptime(fecha_str, "%d/%m/%Y")
        start_str = dt.strftime("%Y%m%dT%H%M%00"); end_str = (dt + timedelta(minutes=30)).strftime("%Y%m%dT%H%M%00")
        titulo = urllib.parse.quote(f"📞 Llamar a {cliente} ({empresa})")
        detalles = urllib.parse.quote(f"CRM Recordatorio de Llamada.\n\nEmpresa: {empresa}\nTeléfono: {telefono}")
        return f"https://calendar.google.com/calendar/render?action=TEMPLATE&text={titulo}&details={detalles}&dates={start_str}/{end_str}"
    except: return ""

# --- MOTOR DE DOCUMENTOS EXCEL ---
FILL_GRIS = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
FILL_BLANCO = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

def procesar_excel(row, obs, tipo_doc):
    try:
        archivo_plantilla = "plantilla_presupuesto.xlsx"
        wb = openpyxl.load_workbook(archivo_plantilla)
        ws = wb.active
    except Exception as e:
        return None, f"🚨 No se encontró '{archivo_plantilla}' en tu carpeta. Asegurate de haberlo guardado ahí."

    try:
        prods_data = json.loads(row['Productos Seleccionados'])
    except:
        prods_data = [{
            "nombre": str(row.get('Productos Seleccionados', 'Cotización General')),
            "desc": "Cotización Manual o antigua",
            "cantidad": "1 pcs.",
            "precio": row['Monto USD / $'],
            "desc_val": "0",
            "importe": row['Monto USD / $']
        }]

    # Inyección de Cabecera
    ws['B10'] = row['Cliente']
    ws['B11'] = row['Empresa']
    ws['B12'] = row['Telefono']
    
    ws['A15'] = row['Asesor']
    ws['B15'] = date.today().strftime("%d/%m/%Y")
    ws['C15'] = row['N° Cotiz.']
    ws['D15'] = "USD" if "USD" in str(row['Monto USD / $']).upper() else "ARS"

    # Inyección de Productos
    fila_inicio = 18
    for i, p in enumerate(prods_data):
        curr_row = fila_inicio + i
        desc_limpia = str(p.get('desc', ''))
        desc_limpia = desc_limpia.replace('•', '   -   ').replace('·', '   -   ')
        
        ws[f'A{curr_row}'] = p['nombre']
        ws[f'B{curr_row}'] = desc_limpia
        ws[f'C{curr_row}'] = p['cantidad']
        ws[f'D{curr_row}'] = p['precio']
        ws[f'E{curr_row}'] = p['desc_val']
        ws[f'F{curr_row}'] = p['importe']

        relleno = FILL_GRIS if i % 2 == 0 else FILL_BLANCO
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{curr_row}'].fill = relleno
            ws[f'{col}{curr_row}'].alignment = Alignment(vertical="top", wrap_text=True)

    # Inyección de Totales y Matemáticas
    monto_base = limpiar_monto_para_suma(row['Monto USD / $'])
    descuento_total = limpiar_monto_para_suma(row.get('Descuento Aplicado', '0'))
    subtotal_bruto = monto_base + descuento_total
    
    if "Argentina" in tipo_doc:
        etiqueta_impuesto = "IVA (21%)"
        impuesto_pct = 0.21
    else:
        etiqueta_impuesto = "Gastos adm. (5%)"
        impuesto_pct = 0.05
        
    ws['E24'] = etiqueta_impuesto

    val_impuestos = monto_base * impuesto_pct
    total_final = monto_base + val_impuestos

    ws['F22'] = subtotal_bruto
    ws['F23'] = descuento_total
    ws['F24'] = val_impuestos
    ws['F25'] = total_final

    # Inyección de Observaciones
    ws['A31'] = f"Observaciones: {obs}"
    
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins.left = 0.2
    ws.page_margins.right = 0.2
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4
    ws.page_setup.orientation = "landscape"
    ws.print_options.horizontalCentered = True
    
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue(), "OK"

@st.cache_data(ttl=30)
def get_data_main():
    try: df = conn.read(worksheet="Central Negociaciones")
    except: df = pd.DataFrame(columns=COLUMNS_MAIN)
    if df is None: df = pd.DataFrame(columns=COLUMNS_MAIN)
    if not df.empty: df.columns = df.columns.astype(str).str.strip()
    for col in COLUMNS_MAIN:
        if col not in df.columns: df[col] = ""
    df['Telefono'] = df['Telefono'].astype(str).str.strip().str.lstrip("'").replace('#ERROR!', '')
    df['N° Cotiz.'] = df['N° Cotiz.'].astype(str).apply(lambda x: x.split('.')[0] if str(x).endswith('.0') else x)
    return df.fillna('')

@st.cache_data(ttl=30)
def get_data_cat():
    try: df = conn.read(worksheet="Catalogo")
    except: df = pd.DataFrame(columns=COLUMNS_CAT)
    if df is None: df = pd.DataFrame(columns=COLUMNS_CAT)
    if not df.empty: df.columns = df.columns.astype(str).str.strip()
    for col in COLUMNS_CAT:
        if col not in df.columns: df[col] = ""
    return df.fillna('')

def guardar_datos(df, sheet="Central Negociaciones"):
    df_safe = df.copy()
    if sheet == "Central Negociaciones":
        df_safe['Telefono'] = df_safe['Telefono'].astype(str).apply(lambda x: f" {x.strip()}" if x.strip().startswith("+") else x.strip())
    conn.update(worksheet=sheet, data=df_safe, spreadsheet=GSHEET_URL)
    st.cache_data.clear()

def generar_numero_cotizacion(df):
    numeros = [int(''.join(filter(str.isdigit, str(val).split('.')[0]))) for val in df['N° Cotiz.'].dropna() if ''.join(filter(str.isdigit, str(val).split('.')[0]))]
    return f"{max(max(numeros) + 1 if numeros else 0, 1000):06d}"

def guardar_gestion(indice, nota_existente, nueva_nota, nueva_fecha_str, fecha_anterior_str):
    df_actual = get_data_main(); fecha_hoy = datetime.now().strftime("%d/%m/%Y")
    texto_agregado = f"[{fecha_hoy}] 📝 {nueva_nota}" if nueva_nota.strip() else ""
    if nueva_fecha_str != str(fecha_anterior_str): texto_agregado += f" | 📅 Reprogramado: {nueva_fecha_str}" if texto_agregado else f"[{fecha_hoy}] 📅 Llamada reprogramada a: {nueva_fecha_str}"
    if texto_agregado: df_actual.at[indice, 'Notas'] = texto_agregado if str(nota_existente).strip() in ["", "nan"] else f"{nota_existente}\n{texto_agregado}"
    df_actual.at[indice, 'Proxima llamada'] = nueva_fecha_str; guardar_datos(df_actual, "Central Negociaciones")

# --- LOGIN ---
if 'autenticado' not in st.session_state:
    st.session_state.autenticado = False; st.session_state.usuario_actual = None

if not st.session_state.autenticado:
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.title("🔐 Acceso CRM DICAD")
        with st.form("login_form"):
            user = st.selectbox("Seleccione su nombre", list(USUARIOS.keys()))
            password = st.text_input("Contraseña", type="password")
            if st.form_submit_button("Ingresar", use_container_width=True):
                if USUARIOS[user] == password:
                    st.session_state.autenticado = True; st.session_state.usuario_actual = user; st.rerun()
                else: st.error("Contraseña incorrecta")
    st.stop()

# --- SIDEBAR ---
with st.sidebar:
    st.markdown('<style>[data-testid="stSidebar"] {background-color: #2E3E57 !important;}</style>', unsafe_allow_html=True)
    st.columns([1, 4, 1])[1].image("logo_dicad.png", use_column_width=True) 
    st.markdown("<p style='text-align: center; color:#fff; font-size:16px; margin-top:0.5em; font-weight: bold;'>CRM DICAD AMÉRICA</p><br>", unsafe_allow_html=True) 
    section = option_menu(None, ["Potenciales", "Pipeline", "Negociaciones", "Agregar Cliente", "Calendario", "Catálogo de Productos"], icons=["person-bounding-box", "kanban", "briefcase", "person-plus", "calendar-date", "box-seam"], default_index=2, styles={"container": {"padding": "5px!important", "background-color": "#F0F2F6", "border-radius": "10px"},"icon": {"color": "#333333", "font-size": "18px"}, "nav-link": {"color": "#333333", "font-size": "16px", "text-align": "left", "margin":"2px 0px", "--hover-color": "#E0E0E0"},"nav-link-selected": {"background-color": "#FF6600", "color": "white"}})
    st.markdown("---")
    st.markdown(f"<div style='text-align: center; color: white; font-size: 14px; margin-bottom: 10px;'>{'👑 Admin' if st.session_state.usuario_actual == ADMINISTRADOR else '💼 Asesor'}: <b>{st.session_state.usuario_actual}</b></div>", unsafe_allow_html=True)
    if st.button("🚪 Cerrar Sesión", use_container_width=True): st.session_state.autenticado = False; st.session_state.usuario_actual = None; st.rerun()

df = get_data_main()
df_cat = get_data_cat()
lista_asesores = ["Todos los Asesores"] + list(USUARIOS.keys())
index_inicio = lista_asesores.index(st.session_state.usuario_actual) if st.session_state.usuario_actual in lista_asesores else 0

# --- CALCULADORA DE PRODUCTOS JSON ---
def modulo_calculadora(key_prefix):
    st.markdown("### 🛒 Configurador de Cotización")
    if df_cat.empty:
        st.warning("⚠️ No hay productos en el catálogo."); return st.text_input("Monto", key=f"mm_{key_prefix}"), "", ""
    
    opciones_base = df_cat['Producto'].tolist()
    opciones_prods = []
    for p in opciones_base:
        opciones_prods.extend([f"{p} (Línea 1)", f"{p} (Línea 2)", f"{p} (Línea 3)", f"{p} (Línea 4)", f"{p} (Línea 5)"])
    
    seleccion = st.multiselect("Seleccioná los softwares a cotizar:", opciones_prods, key=f"sel_{key_prefix}")
    
    subtotal_general = 0.0; total_final = 0.0; ahorro_total = 0.0; moneda_ref = "USD"; json_data = []
    
    if seleccion:
        st.markdown("---")
        for i, s in enumerate(seleccion):
            nombre_real = s.rsplit(' (Línea', 1)[0]
            fila_prod = df_cat[df_cat['Producto'] == nombre_real].iloc[0]
            precio_uni = limpiar_monto_para_suma(fila_prod['Precio'])
            if fila_prod['Moneda']: moneda_ref = str(fila_prod['Moneda']).upper().strip()
            
            st.markdown(f"**📦 {nombre_real}** (Precio Unitario: {moneda_ref} {precio_uni:,.0f})")
            
            c_q, c_d1, c_d2, c_d3 = st.columns([1, 1.5, 1.5, 2])
            cantidad = c_q.number_input("Cantidad", min_value=1, value=1, key=f"cant_{key_prefix}_{i}")
            tipo_desc = c_d1.selectbox("Descuento en:", ["Porcentaje (%)", "Monto Fijo"], key=f"td_{key_prefix}_{i}")
            val_desc = c_d2.number_input("Valor Desc.", min_value=0.0, value=0.0, key=f"vd_{key_prefix}_{i}")
            
            subtotal_linea = precio_uni * cantidad
            
            if "Porcentaje" in tipo_desc: 
                monto_desc = subtotal_linea * (val_desc / 100)
                txt_desc = f"{val_desc}%"
            else: 
                monto_desc = val_desc
                txt_desc = f"{moneda_ref} {val_desc}"
                
            precio_final_prod = subtotal_linea - monto_desc
            c_d3.markdown(f"<div style='margin-top:28px; font-weight:bold; color:#28a745; font-size:15px;'>👉 Total Línea: {moneda_ref} {precio_final_prod:,.0f}</div>", unsafe_allow_html=True)
            
            subtotal_general += subtotal_linea
            total_final += precio_final_prod
            ahorro_total += monto_desc
            
            json_data.append({
                "nombre": nombre_real, "desc": fila_prod['Descripcion'], "cantidad": f"{cantidad} pcs.",
                "precio": f"{moneda_ref} {precio_uni:,.0f}", "desc_val": txt_desc,
                "importe": f"{moneda_ref} {precio_final_prod:,.0f}",
                "raw_precio": precio_uni, "raw_importe": precio_final_prod
            })
            st.markdown("<hr style='margin:10px 0; border: 0; border-top: 1px dashed #ccc;'>", unsafe_allow_html=True)
            
        st.success(f"### 💰 TOTAL FINAL A COBRAR: {moneda_ref} {total_final:,.2f}")
        texto_ahorro = f"Ahorro {moneda_ref} {ahorro_total:,.0f}" if ahorro_total > 0 else "Sin descuento"
        return f"{moneda_ref} {total_final:,.0f}", json.dumps(json_data), texto_ahorro
    else:
        st.info("Seleccioná al menos un producto.")
        return "", "", ""

# --- CATÁLOGO ---
if section == "Catálogo de Productos":
    st.markdown("## 📦 Catálogo Central de Productos")
    if st.session_state.usuario_actual == ADMINISTRADOR:
        with st.expander("➕ Cargar Nuevo Producto al Catálogo", expanded=False):
            with st.form("form_nuevo_prod"):
                c1, c2 = st.columns([2, 3])
                with c1: prod_nombre = st.text_input("Nombre del Producto / Módulo *")
                with c2: prod_desc = st.text_input("Descripción breve (Qué incluye)")
                
                c3, c4, c5 = st.columns(3)
                with c3: prod_cat = st.text_input("Categoría (Ej: Software, Soporte)")
                with c4: prod_moneda = st.selectbox("Moneda", ["USD", "ARS"])
                with c5: prod_precio = st.number_input("Precio de Lista", min_value=0.0)
                
                if st.form_submit_button("Guardar en Catálogo", type="primary"):
                    if not prod_nombre.strip(): st.error("Falta el nombre del producto.")
                    else:
                        nuevo_prod = pd.DataFrame([{"Producto": prod_nombre, "Descripcion": prod_desc, "Categoria": prod_cat, "Moneda": prod_moneda, "Precio": prod_precio}])
                        guardar_datos(pd.concat([df_cat, nuevo_prod], ignore_index=True), "Catalogo")
                        st.success("¡Producto cargado exitosamente!"); st.rerun()
    
    st.markdown("---")
    c_cat1, c_cat2 = st.columns([4,1])
    with c_cat1: st.write("Lista de productos activos")
    with c_cat2: 
        if st.button("🔄 Refrescar Catálogo", use_container_width=True): st.cache_data.clear(); st.rerun()
    st.dataframe(df_cat, use_container_width=True, hide_index=True)

# --- AGREGAR CLIENTE ---
elif section == "Agregar Cliente":
    c_t, c_b = st.columns([4, 1])
    with c_t: st.markdown("## 🙋‍♂️ Nuevo Contacto")
    with c_b: 
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🔄 Refrescar Precios", use_container_width=True): st.cache_data.clear(); st.rerun()

    if 'f_k' not in st.session_state: st.session_state.f_k = 0
    fk = st.session_state.f_k
    
    tipo = st.radio("Fase:", ["🎯 Potencial (Solo contacto)", "💼 Negociación Activa (Cotizar ahora)"], key=f"t_{fk}", horizontal=True); st.markdown("---")
    c1, c2 = st.columns(2)
    with c1:
        cli = st.text_input("Nombre del Cliente *", key=f"c_{fk}"); emp = st.text_input("Empresa", key=f"e_{fk}")
        p_s = st.selectbox("País", CODIGOS_PAISES, key=f"p_{fk}"); ciu = st.text_input("Ciudad", key=f"ci_{fk}")
    with c2:
        prof = st.text_input("Profesión", key=f"pr_{fk}"); car = st.text_input("Cargo", key=f"ca_{fk}")
        tel = st.text_input("Teléfono (Sin código)", key=f"te_{fk}"); eml = st.text_input("Email", key=f"em_{fk}")
    
    st.markdown("---")
    cc1, cc2, cc3 = st.columns([2, 2, 2])
    with cc1: px_l = st.date_input("Próxima llamada", key=f"px_{fk}")
    with cc2: px_h = st.time_input("Hora de llamada", value=datetime.strptime("10:00", "%H:%M").time(), key=f"pxh_{fk}")
    with cc3: cot = st.text_input("N° Cotiz (Vacío=Auto)", key=f"co_{fk}")
    
    ase = st.selectbox("Asesor Asignado", list(USUARIOS.keys()), index=list(USUARIOS.keys()).index(st.session_state.usuario_actual) if st.session_state.usuario_actual in USUARIOS else 0, key=f"as_{fk}") if st.session_state.usuario_actual == ADMINISTRADOR else st.session_state.usuario_actual
    
    monto_final, prods_final, desc_final = "", "", ""
    if "Negociación" in tipo:
        st.markdown("---")
        monto_final, prods_final, desc_final = modulo_calculadora(f"add_{fk}")
        l_p = st.text_input("Link al Presupuesto PDF", key=f"pd_{fk}")
    else: l_p = ""

    st.markdown("---"); n_i = st.text_area("Nota inicial", key=f"ni_{fk}")
    
    if st.button("💾 GUARDAR CLIENTE EN CRM", type="primary", use_container_width=True):
        if not cli.strip(): st.warning("El nombre del cliente es obligatorio.")
        else:
            with st.spinner("Guardando..."):
                p_f, c_f = extraer_pais_codigo(p_s)
                tel_f = f"{c_f} {tel}" if (tel.strip() and not tel.startswith("+") and c_f) else tel
                em_l, te_l = eml.strip().lower(), tel_f.replace(" ","").replace("+","").replace("-","")
                if any((str(r['Email']).lower().strip() == em_l and em_l) or (str(r['Telefono']).replace(" ","").replace("+","").replace("-","") == te_l and te_l) for _, r in df.iterrows()): st.error("🚨 ¡CLIENTE EXISTENTE! Usá 'Nueva Cotización' en la ficha del cliente.")
                else:
                    est_i = "Potencial" if "Potencial" in tipo else "En Proceso"
                    cot_f = cot.strip() if cot.strip() else (generar_numero_cotizacion(df) if est_i == "En Proceso" else "")
                    fh_str = f"{px_l.strftime('%d/%m/%Y')} {px_h.strftime('%H:%M')}"
                    new = pd.DataFrame([{"Creado":datetime.now().strftime("%d/%m/%Y"),"Cliente":cli,"Profesion":prof,"Pais":p_f,"Ciudad":ciu,"Empresa":emp,"Cargo":car,"Telefono":tel_f,"Email":eml,"N° Cotiz.":cot_f,"Monto USD / $":monto_final,"Notas":f"[{datetime.now().strftime('%d/%m/%Y')}] 📝 {n_i}" if n_i else "","Proxima llamada":fh_str,"Asesor":ase,"Estado_Nego":est_i,"Link_PDF":l_p, "Productos Seleccionados":prods_final, "Descuento Aplicado":desc_final}])
                    guardar_datos(pd.concat([df, new], ignore_index=True)); st.session_state.f_k += 1; st.success("Guardado exitosamente!"); st.rerun()

# --- NEGOCIACIONES ---
elif section == "Negociaciones":
    st.markdown("## :card_index_dividers: Negociaciones Activas")
    df_nego = df[(df['Estado_Nego'] != 'Potencial') & (df['Estado_Nego'] != '')]

    if st.session_state.usuario_actual == ADMINISTRADOR:
        st.markdown("### 🏢 PANEL ESTRATÉGICO (Admin)")
        c1, c2, c3 = st.columns(3)
        c1.metric("💰 Total Empresa (USD)", f"USD {sum(limpiar_monto_para_suma(x) for x in df_nego['Monto USD / $'] if 'ARS' not in str(x).upper()):,.0f}")
        c2.metric("💵 Total Empresa (ARS)", f"ARS {sum(limpiar_monto_para_suma(x) for x in df_nego['Monto USD / $'] if 'ARS' in str(x).upper()):,.0f}")
        c3.metric("🤝 Negociaciones Totales", len(df_nego))
        st.markdown("---")

    asesor_sel = st.selectbox("Seleccionar Asesor:", lista_asesores, index=index_inicio)
    df_tab = df_nego if asesor_sel == "Todos los Asesores" else df_nego[df_nego['Asesor'] == asesor_sel]

    c_b, c_r = st.columns([4, 1])
    with c_b: busq = st.text_input("🔍 Buscar Cliente/Empresa:")
    with c_r: 
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🔄 Actualizar Datos"): st.cache_data.clear(); st.rerun()
                
    df_f = df_tab[df_tab['Cliente'].astype(str).str.contains(busq, case=False) | df_tab['Empresa'].astype(str).str.contains(busq, case=False)] if busq else df_tab

    for idx, row in df_f.iterrows():
        est = row.get('Estado_Nego', 'En Proceso'); color = "#28a745" if est == 'Ganada' else "#dc3545" if est == 'Perdida' else "#ffc107"
        prox_llamada = row.get('Proxima llamada', '')
        
        try:
            prods_json = json.loads(row.get('Productos Seleccionados', '[]'))
            nombres_p = [p['nombre'] for p in prods_json]
            texto_prods = " + ".join(nombres_p)
        except: texto_prods = str(row.get('Productos Seleccionados', ''))
        
        prod_badge = f"<br><small style='color:#555;'>📦 <b>Incluye:</b> {texto_prods}</small>" if texto_prods else ""
        desc_badge = f" | <small style='color:#dc3545;'>{row.get('Descuento Aplicado', '')}</small>" if row.get('Descuento Aplicado') and row.get('Descuento Aplicado') != "Sin descuento" else ""
        
        badge_fecha = f"<br><span style='background:#6c757d;color:white;padding:4px 8px;border-radius:6px;font-size:11px;font-weight:bold; display:inline-block; margin-top:5px;'>📅 {prox_llamada}</span>" if str(prox_llamada).strip() else ""
        
        html_card = f"<div style='background:white;padding:1.3em;border-radius:12px;margin-bottom:0.6em;box-shadow:0 1px 8px #d0d6e1;border-left:6px solid {color};color:black; overflow:hidden;'><div style='float:right; text-align:right;'><span style='background:{color};color:white;padding:4px 8px;border-radius:6px;font-size:12px;font-weight:bold; display:inline-block;'>{'✅' if est=='Ganada' else '❌' if est=='Perdida' else '⏳'} {est.upper()}</span>{badge_fecha}</div><b>Cliente:</b> {row.get('Cliente', '')} | <b>Cotiz:</b> {row.get('N° Cotiz.', 'N/A')}<br><b>Monto Final:</b> <span style='color:#2261b6;font-weight:bold;font-size:16px;'>{row.get('Monto USD / $', '')}</span>{desc_badge}{prod_badge}</div>"
        st.markdown(html_card, unsafe_allow_html=True)
        
        with st.expander("📞 ASISTENTE DE LLAMADA (Manejo de Objeciones de Cierre)", expanded=False):
            st.warning("🎯 **Modo Cierre Activado:** Aislá la objeción antes de responder o ceder precio.")
            st.markdown("🛡️ **Si dicen:** *'Llamame la semana que viene'* <br> **Tu respuesta:** *'Entiendo {}. Exactamente, ¿qué va a cambiar de esta semana a la próxima que haga diferente la decisión?'*".format(row.get('Cliente','').split(' ')[0]), unsafe_allow_html=True)
            st.markdown("🛡️ **Si dicen:** *'Está muy caro / Se va de presupuesto'* <br> **Tu respuesta:** *'Aparte del precio, ¿hay alguna otra cosa que te impida avanzar hoy con nosotros?'*", unsafe_allow_html=True)
            st.markdown("🛡️ **Si dicen:** *'Lo tengo que consultar con mi socio'* <br> **Tu respuesta:** *'Perfecto. Y vos personalmente, ¿qué le vas a recomendar a tu socio que hagan?'*", unsafe_allow_html=True)

        with st.expander("Ver Ficha Completa"):
            puede = (st.session_state.usuario_actual == ADMINISTRADOR) or (st.session_state.usuario_actual == row.get('Asesor', ''))
            
            if puede:
                # --- NUEVO MOTOR DE DOCUMENTOS EXCEL ---
                st.markdown("### 📄 Generar Presupuesto Excel")
                c_wd1, c_wd2 = st.columns([2, 1])
                with c_wd1: obs_excel = st.text_area("Observaciones para el cliente:", key=f"obs_e_{idx}")
                with c_wd2: tipo_plantilla = st.radio("Tipo de Presupuesto:", ["Argentina (IVA 21%)", "Internacional (Gasto Adm 5%)"], key=f"tpl_{idx}")
                
                if st.button("⚙️ Procesar Archivo Excel", key=f"btn_e_{idx}"):
                    data_file, status = procesar_excel(row, obs_excel, tipo_plantilla)
                    if data_file: st.session_state[f"doc_ready_{idx}"] = data_file
                    else: st.error(status)
                
                if st.session_state.get(f"doc_ready_{idx}"):
                    st.download_button(label="⬇️ Descargar Presupuesto Listo (.xlsx)", data=st.session_state[f"doc_ready_{idx}"], file_name=f"Presupuesto_{row['Cliente'].replace(' ','_')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"dl_{idx}")
                st.markdown("---")

                with st.expander("⚙️ Editar Datos del Contacto / Cotización"):
                    c_e1, c_e2 = st.columns(2)
                    with c_e1: 
                        ec = st.text_input("Nombre", row.get('Cliente',''), key=f"ecn_{idx}")
                        ee = st.text_input("Empresa", row.get('Empresa',''), key=f"een_{idx}")
                        e_prof = st.text_input("Profesión", row.get('Profesion',''), key=f"eprn_{idx}")
                        e_cargo = st.text_input("Cargo", row.get('Cargo',''), key=f"ecrn_{idx}")
                        edm = st.text_input("Corrección Manual Monto Final", row.get('Monto USD / $',''), key=f"edm_{idx}")
                    with c_e2:
                        idx_pa = next((i for i, p in enumerate(CODIGOS_PAISES) if str(row.get('Pais','')).lower() in p.lower() and row.get('Pais','') != ""), 0)
                        ep = st.selectbox("País", CODIGOS_PAISES, index=idx_pa, key=f"epn_{idx}")
                        e_ciu = st.text_input("Ciudad", row.get('Ciudad',''), key=f"eciun_{idx}")
                        em = st.text_input("Email", row.get('Email',''), key=f"emn_{idx}")
                        etel = st.text_input("Teléfono", row.get('Telefono',''), key=f"eteln_{idx}")
                        edl = st.text_input("Corrección Link PDF", row.get('Link_PDF',''), key=f"edl_{idx}")
                    if st.button("💾 Actualizar Todo", key=f"becn_{idx}", type="primary"):
                        p_n, c_n = extraer_pais_codigo(ep)
                        tel_f = f"{c_n} {etel}" if (etel.strip() and not etel.startswith("+") and c_n) else etel
                        df_u = get_data_main()
                        df_u.loc[idx, ['Cliente','Empresa','Profesion','Cargo','Pais','Ciudad','Email','Telefono', 'Monto USD / $', 'Link_PDF']] = [ec, ee, e_prof, e_cargo, p_n, e_ciu, em, tel_f, edm, edl]
                        guardar_datos(df_u); st.rerun()

            col1, col2 = st.columns(2)
            with col1: st.write(f"**Prof:** {row.get('Profesion','')} | **Cargo:** {row.get('Cargo','')}"); st.write(f"**Tel:** {row.get('Telefono','')}"); st.write(f"**Email:** {row.get('Email','')}")
            with col2: st.write(f"**Empresa:** {row.get('Empresa','')}"); st.write(f"**Cotiz:** {row.get('N° Cotiz.','')}"); st.write(f"**Asesor:** {row.get('Asesor','')}")
            if row.get('Link_PDF'): st.link_button("📄 Ver Presupuesto Subido (Nube)", row['Link_PDF'], use_container_width=True)

            if puede:
                # --- NUEVA LÓGICA: ACTUALIZAR COTIZACIÓN EN LA MISMA NEGOCIACIÓN ---
                st.markdown("---"); st.markdown("### 🔄 Generar Nueva Cotización (Misma Negociación)")
                st.info("Al crear una cotización nueva, se reemplazará el monto y productos de la cotización actual, pero la negociación y el historial seguirán intactos.")
                m_f, p_f, d_f = modulo_calculadora(f"nc_{idx}")
                c_nc1, c_nc2 = st.columns(2)
                with c_nc1: ncc = st.text_input("N° Cotiz (Vacío=Auto)", key=f"ncc_{idx}")
                with c_nc2: ncp = st.text_input("Link al PDF en la nube (Opcional)", key=f"ncp_{idx}")
                
                col_btn1, col_btn2 = st.columns(2)
                with col_btn1:
                    if st.button("🚀 Actualizar Cotización en Nego", key=f"bnc_{idx}", type="primary", use_container_width=True):
                        df_n = get_data_main()
                        f_h = datetime.now().strftime("%d/%m/%Y")
                        num_c = ncc if ncc else generar_numero_cotizacion(df_n)
                        nota_previa = str(df_n.at[idx, 'Notas'])
                        
                        df_n.at[idx, 'N° Cotiz.'] = num_c
                        df_n.at[idx, 'Monto USD / $'] = m_f
                        df_n.at[idx, 'Link_PDF'] = ncp
                        df_n.at[idx, 'Productos Seleccionados'] = p_f
                        df_n.at[idx, 'Descuento Aplicado'] = d_f
                        df_n.at[idx, 'Notas'] = nota_previa + f"\n[{f_h}] 🔄 Nueva Cotización {num_c} generada por {m_f}."
                        guardar_datos(df_n); st.rerun()
                with col_btn2:
                    if st.button("🗑️ Descartar Cotiz. Actual (Sigue en Nego)", key=f"desc_{idx}", use_container_width=True):
                        df_n = get_data_main()
                        f_h = datetime.now().strftime("%d/%m/%Y")
                        nota_previa = str(df_n.at[idx, 'Notas'])
                        df_n.at[idx, 'N° Cotiz.'] = "Rechazada"
                        df_n.at[idx, 'Monto USD / $'] = "0"
                        df_n.at[idx, 'Notas'] = nota_previa + f"\n[{f_h}] 📉 Cotización anterior descartada por el cliente. Negociación sigue abierta."
                        guardar_datos(df_n); st.rerun()

            st.markdown("---"); st.markdown("**📝 Seguimiento y Gestión:**"); st.caption(row.get('Notas',''))
            if puede:
                cn1, cn2, cn3 = st.columns([1.5, 2, 1.5])
                with cn1:
                    f_o, h_o = parsear_fecha_hora(row.get('Proxima llamada', ''))
                    cc1, cc2 = st.columns(2)
                    with cc1: nf = st.date_input("Día", value=f_o, key=f"fgn_{idx}")
                    with cc2: nh = st.time_input("Hora", value=h_o, key=f"hgn_{idx}")
                with cn2: nn = st.text_input("Nueva nota para historial", key=f"ngn_{idx}")
                with cn3: 
                    st.markdown("<br>", unsafe_allow_html=True)
                    if st.button("💾 Guardar Gestión", key=f"bgn_{idx}", use_container_width=True): 
                        fh_str = f"{nf.strftime('%d/%m/%Y')} {nh.strftime('%H:%M')}"
                        guardar_gestion(idx, row.get('Notas',''), nn, fh_str, row.get('Proxima llamada','')); st.rerun()
                    if st.button("✅ Llamada OK", key=f"ok_nego_{idx}", use_container_width=True):
                        df_u = get_data_main()
                        nota_previa = str(df_u.at[idx, 'Notas'])
                        df_u.at[idx, 'Notas'] = nota_previa + f"\n[{datetime.now().strftime('%d/%m/%Y')}] ✅ Llamada completada."
                        df_u.at[idx, 'Proxima llamada'] = ""
                        guardar_datos(df_u); st.rerun()

                st.markdown("---")
                if est in ['Ganada','Perdida']:
                    if st.button("🔄 Reabrir Negociación", key=f"re_{idx}"): df_r = get_data_main(); df_r.at[idx, 'Estado_Nego'] = "En Proceso"; guardar_datos(df_r); st.rerun()
                else:
                    cg, cp = st.columns(2)
                    with cg: 
                        if st.button("✅ CERRADA GANADA", key=f"g_{idx}", use_container_width=True): df_g = get_data_main(); df_g.at[idx, 'Estado_Nego'] = "Ganada"; df_g.at[idx, 'Notas'] = str(df_g.at[idx,'Notas']) + f"\n[{datetime.now().strftime('%d/%m/%Y')}] 🏆 NEGOCIACIÓN CERRADA Y GANADA"; guardar_datos(df_g); st.rerun()
                    with cp: 
                        if st.button("❌ CERRADA PERDIDA", key=f"p_{idx}", use_container_width=True): df_p = get_data_main(); df_p.at[idx, 'Estado_Nego'] = "Perdida"; df_p.at[idx, 'Notas'] = str(df_p.at[idx,'Notas']) + f"\n[{datetime.now().strftime('%d/%m/%Y')}] ❌ NEGOCIACIÓN CERRADA Y PERDIDA"; guardar_datos(df_p); st.rerun()

# --- POTENCIALES ---
elif section == "Potenciales":
    st.markdown("## 🎯 Clientes Potenciales")
    asesor_sel = st.selectbox("Filtrar por Asesor:", lista_asesores, index=index_inicio)
    df_pot = df[df['Estado_Nego'] == 'Potencial']
    if asesor_sel != "Todos los Asesores": df_pot = df_pot[df_pot['Asesor'] == asesor_sel]
    st.metric("Total de Leads", len(df_pot)); st.markdown("---")

    for idx, row in df_pot.iterrows():
        puede_editar = (st.session_state.usuario_actual == ADMINISTRADOR) or (st.session_state.usuario_actual == row.get('Asesor', ''))
        st.markdown(f'<div style="background:white;padding:1em;border-radius:10px;border-left:5px solid #6c757d;margin-bottom:0.5em;box-shadow:0 1px 4px #d0d6e1;color:black;"><b>Cliente:</b> {row.get("Cliente", "")} | <b>Empresa:</b> {row.get("Empresa", "")} <br> <b>Teléfono:</b> {row.get("Telefono", "")} | <b>Próx:</b> {row.get("Proxima llamada", "")}</div>', unsafe_allow_html=True)
        
        with st.expander("📞 ASISTENTE DE LLAMADA (Guiones de Descubrimiento)", expanded=False):
            st.warning("🗣️ **Objetivo de esta llamada:** Descubrir el dolor del cliente y generar interés para enviar presupuesto.")
            st.markdown("💡 **Tip 1:** ¿Qué desafío estructural o de tiempos los motivó a buscar nuevas herramientas?")
            st.markdown("💡 **Tip 2:** ¿Qué software están usando hoy y qué es lo que más les frustra de ese proceso?")
        
        with st.expander(f"Ver / Editar a {row.get('Cliente', '')}"):
            st.info(f"**Historial:**\n{row.get('Notas', 'Sin notas.')}")
            if puede_editar:
                c_n1, c_n2, c_n3 = st.columns([1.5, 2, 1.5])
                with c_n1:
                    f_o, h_o = parsear_fecha_hora(row.get('Proxima llamada', ''))
                    cc1, cc2 = st.columns(2)
                    with cc1: n_f = st.date_input("Día", value=f_o, key=f"f_p_{idx}")
                    with cc2: n_h = st.time_input("Hora", value=h_o, key=f"h_p_{idx}")
                with c_n2: n_n = st.text_input("Nota hoy", key=f"n_p_{idx}")
                with c_n3: 
                    st.markdown("<br>", unsafe_allow_html=True)
                    if st.button("💾 Guardar", key=f"b_p_{idx}"):
                        fh_str = f"{n_f.strftime('%d/%m/%Y')} {n_h.strftime('%H:%M')}"
                        guardar_gestion(idx, row.get('Notas',''), n_n, fh_str, row.get('Proxima llamada','')); st.rerun()
                    if st.button("✅ Llamada OK", key=f"ok_pot_{idx}", use_container_width=True):
                        df_u = get_data_main()
                        nota_previa = str(df_u.at[idx, 'Notas'])
                        df_u.at[idx, 'Notas'] = nota_previa + f"\n[{datetime.now().strftime('%d/%m/%Y')}] ✅ Llamada completada."
                        df_u.at[idx, 'Proxima llamada'] = ""
                        guardar_datos(df_u); st.rerun()
                st.markdown("---")
                
                st.markdown("### 🚀 Promover a Negociación")
                m_f, p_f, d_f = modulo_calculadora(f"prov_{idx}")
                n_l_p = st.text_input("Link PDF (Opcional)", key=f"pl_{idx}")
                if st.button("🚀 Guardar Cotización y Promover", type="primary"):
                    df_a = get_data_main(); df_a.at[idx, 'Estado_Nego'] = "En Proceso"
                    df_a.at[idx, 'Monto USD / $'] = m_f; df_a.at[idx, 'Link_PDF'] = n_l_p
                    df_a.at[idx, 'Productos Seleccionados'] = p_f; df_a.at[idx, 'Descuento Aplicado'] = d_f
                    if not str(df_a.at[idx, 'N° Cotiz.']).strip(): df_a.at[idx, 'N° Cotiz.'] = generar_numero_cotizacion(df_a)
                    guardar_datos(df_a); st.rerun()

# --- PIPELINE KANBAN ---
elif section == "Pipeline":
    c_t, c_b = st.columns([4, 1])
    with c_t: st.markdown("## 📊 Pipeline de Ventas")
    with c_b: 
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🔄 Actualizar Tablero", use_container_width=True): st.cache_data.clear(); st.rerun()
        
    asesor_sel = st.selectbox("Filtrar Tablero por Asesor:", lista_asesores, index=index_inicio)
    df_pipe = df if asesor_sel == "Todos los Asesores" else df[df['Asesor'] == asesor_sel]
    estados_kanban = ["Potencial", "En Proceso", "Ganada", "Perdida"]; cols_kanban = st.columns(4)
    
    for i, estado in enumerate(estados_kanban):
        with cols_kanban[i]:
            df_col = df_pipe[df_pipe['Estado_Nego'] == estado]
            tot_usd = sum(limpiar_monto_para_suma(x) for x in df_col['Monto USD / $'] if 'ARS' not in str(x).upper())
            color_header = "#6c757d" if estado=="Potencial" else "#ffc107" if estado=="En Proceso" else "#28a745" if estado=="Ganada" else "#dc3545"
            st.markdown(f"<div style='background-color:{color_header}; color:white; padding:10px; border-radius:5px; text-align:center; font-weight:bold; margin-bottom:15px;'>{estado.upper()} ({len(df_col)})<br>USD {tot_usd:,.0f}</div>", unsafe_allow_html=True)
            
            for idx, row in df_col.iterrows():
                puede = (st.session_state.usuario_actual == ADMINISTRADOR) or (st.session_state.usuario_actual == row.get('Asesor', ''))
                
                st.markdown(f"<div style='background:white; padding:12px; border-radius:8px; box-shadow:0 2px 5px rgba(0,0,0,0.15); margin-bottom:5px; border-left:4px solid {color_header}; color:black;'><b style='font-size:14px;'>{row.get('Cliente','')}</b><br><span style='font-size:12px; color:#555;'>{row.get('Empresa','')}</span><br><b style='font-size:13px; color:#2261b6;'>{row.get('Monto USD / $','')}</b><br><span style='font-size:11px; color:#888;'>📅 {row.get('Proxima llamada','')}</span></div>", unsafe_allow_html=True)
                
                if puede:
                    opciones_mover = ["Mover a..."] + [e for e in estados_kanban if e != estado]
                    nuevo_est = st.selectbox("Acción", opciones_mover, key=f"mov_{idx}", label_visibility="collapsed")
                    if nuevo_est != "Mover a...":
                        df_actual = get_data_main()
                        df_actual.at[idx, 'Estado_Nego'] = nuevo_est
                        if nuevo_est == "En Proceso" and not str(df_actual.at[idx, 'N° Cotiz.']).strip(): df_actual.at[idx, 'N° Cotiz.'] = generar_numero_cotizacion(df_actual)
                        fecha_hoy = datetime.now().strftime("%d/%m/%Y"); nota_cambio = f"[{fecha_hoy}] 🔄 Movido a: {nuevo_est.upper()}"; nota_previa = str(df_actual.at[idx,'Notas'])
                        df_actual.at[idx, 'Notas'] = nota_cambio if nota_previa.strip() in ["", "nan"] else f"{nota_previa}\n{nota_cambio}"
                        guardar_datos(df_actual); st.rerun()

# --- CALENDARIO ---
elif section == "Calendario":
    c_t, c_b = st.columns([4, 1])
    with c_t: st.markdown("## 📅 Agenda Diaria")
    with c_b: 
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🔄 Actualizar", use_container_width=True): st.cache_data.clear(); st.rerun()
    
    # Filtramos para que NO muestre los que tienen la fecha vacía (llamadas realizadas)
    df_a = df[df['Estado_Nego'].isin(['En Proceso', 'Potencial']) & (df['Proxima llamada'].astype(str).str.strip() != '')].copy()
    
    if df_a.empty: st.success("Todo al día. No hay llamadas pendientes.")
    else:
        df_a['F'] = pd.to_datetime(df_a['Proxima llamada'], format='%d/%m/%Y %H:%M', errors='coerce').fillna(pd.to_datetime(df_a['Proxima llamada'], format='%d/%m/%Y', errors='coerce'))
        for idx, r in df_a.sort_values('F').iterrows():
            link_cal = generar_link_gcal(r['Cliente'], r['Empresa'], r['Telefono'], r['Proxima llamada'])
            btn_cal = f"<a href='{link_cal}' target='_blank' style='text-decoration:none; background-color:#4285F4; color:white; padding:4px 8px; border-radius:4px; font-size:12px; font-weight:bold; margin-left:10px;'>📅 Añadir a Google Calendar</a>" if link_cal else ""
            st.markdown(f'<div style="background:#2E3E57;padding:15px;border-radius:10px;margin-bottom:10px;border-left:5px solid #FF6600;"><h4 style="color:white;margin:0;">📅 {r["Proxima llamada"]} hs | {r["Cliente"]} <span style="font-size:12px;background:#556B8D;padding:3px 8px;border-radius:5px;margin-left:5px;">{"🎯" if r["Estado_Nego"]=="Potencial" else "💼"}</span> {btn_cal}</h4><p style="color:#d0d6e1;margin:5px 0;">📞 {r["Telefono"]} | ✉️ {r["Email"]} | 👔 {r["Asesor"]}</p></div>', unsafe_allow_html=True)
            
            if (st.session_state.usuario_actual == ADMINISTRADOR) or (st.session_state.usuario_actual == r.get('Asesor', '')):
                col_btn1, col_btn2 = st.columns(2)
                with col_btn1:
                    with st.expander("⚙️ Reprogramar / Editar Contacto"):
                        c_f1, c_f2 = st.columns(2)
                        f_val, h_val = parsear_fecha_hora(r.get('Proxima llamada', ''))
                        with c_f1: e_f = st.date_input("Reprogramar Día", value=f_val, key=f"ef_{idx}")
                        with c_f2: e_h = st.time_input("Reprogramar Hora", value=h_val, key=f"eh_{idx}")
                        if st.button("💾 Guardar Reprogramación", key=f"be_{idx}"):
                            fh_str = f"{e_f.strftime('%d/%m/%Y')} {e_h.strftime('%H:%M')}"
                            df_u = get_data_main(); df_u.at[idx, 'Proxima llamada'] = fh_str; guardar_datos(df_u); st.rerun()
                with col_btn2:
                    if st.button("✅ Marcar Llamada Realizada", key=f"ok_cal_{idx}", use_container_width=True):
                        df_u = get_data_main()
                        nota_previa = str(df_u.at[idx, 'Notas'])
                        df_u.at[idx, 'Notas'] = nota_previa + f"\n[{datetime.now().strftime('%d/%m/%Y')}] ✅ Llamada programada completada."
                        df_u.at[idx, 'Proxima llamada'] = ""
                        guardar_datos(df_u); st.rerun()